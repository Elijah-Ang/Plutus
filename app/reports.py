from __future__ import annotations

from datetime import date
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

from .storage import Storage
from .utils import PROJECT_ROOT, json_dumps, redact

SHEETS: list[tuple[str, str | None]] = [
    ("Summary Dashboard", None), ("Proposals", "SELECT *, CASE WHEN status='pending' AND expires_at > strftime('%Y-%m-%dT%H:%M:%f+00:00','now') THEN 'active_pending' WHEN status='pending' THEN 'pending_expired' WHEN status='approved' AND expires_at > strftime('%Y-%m-%dT%H:%M:%f+00:00','now') THEN 'approved_unsubmitted' WHEN status='approved' THEN 'approved_historical' WHEN status='submitted' THEN 'submitted_historical' WHEN status='filled' THEN 'filled_historical' WHEN status='rejected' THEN 'rejected_historical' WHEN status='superseded' THEN 'superseded_historical' WHEN status='blocked' THEN 'blocked_historical' WHEN status='expired' THEN 'expired_historical' ELSE COALESCE(status, 'unknown') END AS proposal_runtime_state, CASE WHEN status='pending' AND expires_at > strftime('%Y-%m-%dT%H:%M:%f+00:00','now') THEN 1 ELSE 0 END AS active_pending_actionable FROM trade_proposals ORDER BY created_at DESC"), ("Daily PnL", "daily_summaries"), ("Trades", "orders"),
    ("Orders", "orders"), ("Fills", "fills"), ("Positions", "positions"), ("Signals", "signals"),
    ("Risk Checks", "risk_checks"), ("AI Reviews", "ai_reviews"), ("Approvals", "approvals"),
    ("Cash Management", "cashout_suggestions"), ("ML Shadow Metrics", "model_versions"),
    ("Market Memory", "market_memory"),
    ("Telegram Digests", "telegram_digests"),
    ("Errors", "errors"), ("Audit Events", "audit_events"), ("Config Snapshot", "config_snapshots"),
    ("Control State", "control_state"),
    ("Sleep Mode Events", "SELECT * FROM audit_events WHERE event_type IN ('sleep_mode_enabled', 'sleep_mode_disabled', 'sleep_mode_command_ignored_unauthorized', 'sleep_mode_command_ignored_old')"),
    ("Emergency Exit Events", "SELECT * FROM audit_events WHERE event_type IN ('emergency_exit_auto_timeout_reached', 'emergency_exit_submitted', 'emergency_exit_blocked', 'emergency_exit_cancelled_by_user', 'emergency_exit_approved_by_user')"),
    ("Emergency Exit Score Breakdown", "SELECT symbol, position_drawdown_pct, average_entry_price, latest_position_price, atr_value, adverse_move_atr, minutes_to_close, emergency_exit_score, emergency_exit_triggered, emergency_exit_trigger_reason, created_at FROM market_memory WHERE emergency_exit_score IS NOT NULL"),
    ("Exit Blocker State", "SELECT symbol, side, status, CASE WHEN status IN ('pending','approved') AND expires_at > strftime('%Y-%m-%dT%H:%M:%f+00:00','now') THEN 'active' ELSE 'stale' END AS blocker_state, CASE WHEN status IN ('pending','approved') AND expires_at > strftime('%Y-%m-%dT%H:%M:%f+00:00','now') AND emergency_exit_triggered=1 THEN symbol || ' emergency exit review active' WHEN status IN ('pending','approved') AND expires_at > strftime('%Y-%m-%dT%H:%M:%f+00:00','now') THEN symbol || ' EXIT proposal pending' ELSE 'stale ' || symbol || ' exit flag ignored' END AS blocker_reason, created_at, expires_at, exit_trigger_reason, emergency_exit_score, emergency_exit_triggered, emergency_exit_trigger_reason FROM trade_proposals WHERE side='sell' AND status IN ('pending','approved','submitted','filled','blocked','expired','rejected','superseded','stale_resolved') ORDER BY created_at DESC"),
    ("Exit Blocked BUY Candidates", "SELECT symbol, price, signal, score, no_action_reason, candidate_suppression_reason, exit_priority_applied, CASE WHEN EXISTS (SELECT 1 FROM trade_proposals p WHERE p.side='sell' AND p.status IN ('pending','approved') AND p.expires_at > strftime('%Y-%m-%dT%H:%M:%f+00:00','now')) THEN 'active_exit_blocker_currently_exists' ELSE 'stale_or_historical_exit_blocker_reason' END AS current_exit_blocker_state, created_at FROM market_memory WHERE signal='ENTRY' AND lower(no_action_reason) LIKE '%exit%'"),
    ("Exit Review Status", "exit_review_events"),
    ("Exit Candidates", "SELECT * FROM exit_review_events WHERE status IN ('exit_candidate','exit_review_needed') ORDER BY created_at DESC"),
    ("Exit Block Reasons", "SELECT symbol, review_type, status, reason, drawdown_from_entry_pct, drawdown_from_peak_pct, created_at FROM exit_review_events WHERE status!='exit_candidate' OR COALESCE(reason,'')!='' ORDER BY created_at DESC"),
    ("Trailing Stop State", "SELECT symbol,current_price,avg_entry_price,highest_price_since_entry,max_unrealized_profit_pct,pullback_from_peak_pct,drawdown_from_peak_pct,trailing_stop_price,decision_type,action,reason,created_at FROM position_management_decisions ORDER BY created_at DESC"),
    ("Time Stop State", "SELECT symbol,review_type,status,reason,time_stop_status,position_age_days,position_age_cycles,unrealized_pl_pct,peak_unrealized_pct,created_at FROM exit_review_events WHERE review_type='TIME_STOP_EXIT' OR time_stop_status='triggered' ORDER BY created_at DESC"),
    ("Position Drawdown Metrics", "SELECT symbol,current_price,avg_entry_price,quantity,unrealized_profit_pct,drawdown_from_entry_pct,drawdown_from_peak_pct,position_age_days,position_age_cycles,decision_type,reason,created_at FROM position_management_decisions ORDER BY created_at DESC"),
    ("Position Peak Giveback Metrics", "SELECT symbol,current_price,highest_price_since_entry,max_unrealized_profit_pct,pullback_from_peak_pct,profit_giveback_ratio,trailing_stop_price,decision_type,reason,created_at FROM position_management_decisions ORDER BY created_at DESC"),
    ("Suppressed Sleep BUY Candidates", "SELECT symbol, price, signal, score, no_action_reason, candidate_suppression_reason, created_at FROM market_memory WHERE candidate_suppression_reason = 'suppressed_by_sleep_mode'"),
    ("Wake Summary Events", "SELECT * FROM audit_events WHERE event_type = 'wake_summary_sent'"),
    ("Performance Lab Summary", "performance_lab_summaries"),
    ("Crypto Research Summary", "SELECT symbol,lane,score,data_freshness,price,price_timestamp,return_1h,return_4h,return_1d,return_7d,return_20d,provider,created_at,'research-only; proposals/orders disabled by default' AS safety_status FROM crypto_research_snapshots ORDER BY created_at DESC, symbol"),
    ("Crypto Candidate Briefs", "SELECT symbol,lane,score,price,data_freshness,trend_metrics,score_components,risk_metrics,'research_only' AS action_mode,'no proposals/orders unless explicitly enabled later' AS blocked_actions,created_at FROM crypto_research_snapshots ORDER BY created_at DESC, score DESC"),
    ("Crypto Observation State", "crypto_observation_state"),
    ("Crypto Counterfactual Outcomes", "crypto_counterfactual_outcomes"),
    ("Crypto Data Coverage", "SELECT symbol,provider,data_freshness,price_timestamp,CASE WHEN price IS NOT NULL THEN 1 ELSE 0 END AS price_available,CASE WHEN return_1h IS NOT NULL THEN 1 ELSE 0 END AS return_1h_available,CASE WHEN return_4h IS NOT NULL THEN 1 ELSE 0 END AS return_4h_available,CASE WHEN return_1d IS NOT NULL THEN 1 ELSE 0 END AS return_1d_available,CASE WHEN return_7d IS NOT NULL THEN 1 ELSE 0 END AS return_7d_available,CASE WHEN return_20d IS NOT NULL THEN 1 ELSE 0 END AS return_20d_available,CASE WHEN realized_volatility IS NOT NULL THEN 1 ELSE 0 END AS realized_volatility_available,CASE WHEN atr_like_volatility IS NOT NULL THEN 1 ELSE 0 END AS atr_like_volatility_available,CASE WHEN volume IS NOT NULL THEN 1 ELSE 0 END AS volume_available,CASE WHEN spread IS NOT NULL THEN 1 ELSE 0 END AS spread_available,created_at FROM crypto_research_snapshots ORDER BY created_at DESC, symbol"),
    ("Crypto Risk Metrics", "SELECT symbol,score,data_freshness,realized_volatility,atr_like_volatility,spread,risk_metrics,'margin disabled' AS margin_status,'shorting disabled' AS shorting_status,'equity sizing/exposure unaffected' AS equity_isolation,created_at FROM crypto_research_snapshots ORDER BY created_at DESC, symbol"),
    ("Crypto Capability", "SELECT id,run_id,provider,broker,trading_mode,market_profile,data_feed,asset_count,official_contract_version,official_contract_fingerprint,config_hash,formula_version,schema_version,captured_at,expires_at,authoritative,failure_reasons_json,snapshot_fingerprint FROM crypto_capability_snapshots ORDER BY captured_at DESC"),
    ("Crypto Pair Precision", "SELECT snapshot_id,symbol,asset_class,exchange,status,tradable,fractionable,marginable,shortable,easy_to_borrow,min_order_size,min_trade_increment,price_increment,base_asset,quote_currency,authoritative,failure_reasons_json,asset_fingerprint FROM crypto_asset_capabilities ORDER BY snapshot_id DESC,symbol"),
    ("Crypto Market Evidence", "SELECT id,run_id,research_run_id,capability_snapshot_id,symbol,provider,data_feed,bid_price,ask_price,bid_size,ask_size,quote_timestamp,quote_age_seconds,trade_price,trade_size,trade_timestamp,trade_age_seconds,orderbook_bid_price,orderbook_ask_price,orderbook_bid_size,orderbook_ask_size,orderbook_timestamp,orderbook_age_seconds,spread_bps,top_of_book_notional,authoritative,execution_eligible,failure_reasons_json,warnings_json,config_hash,formula_version,schema_version,captured_at,evidence_fingerprint FROM crypto_market_data_evidence ORDER BY captured_at DESC,symbol"),
    ("Crypto Portfolio Risk Evidence", "SELECT id,run_id,symbol,paper_account_id_hash,capability_snapshot_id,market_evidence_id,aggregate_json,derived_authority_json,authoritative,failure_reasons_json,config_hash,formula_version,schema_version,captured_at,expires_at,snapshot_fingerprint FROM crypto_risk_snapshots ORDER BY captured_at DESC,symbol"),
    ("Crypto Decimal Sizing", "SELECT id,run_id,symbol,side,action,request_basis,limit_price,stop_price,stop_execution_price,canonical_quantity,canonical_notional,canonical_stop_risk,gross_stop_risk,estimated_fees,estimated_stop_slippage,minimum_order_size,quantity_increment,price_increment,eligible,authoritative,execution_authorized,blockers_json,binding_caps_json,risk_snapshot_id,config_hash,formula_version,schema_version,created_at,decision_fingerprint FROM crypto_sizing_decisions ORDER BY created_at DESC,symbol"),
    ("Crypto Risk Decisions", "SELECT id,run_id,snapshot_id,sizing_decision_id,risk_eligible,execution_authorized,checks_json,reasons_json,config_hash,formula_version,schema_version,created_at,decision_fingerprint FROM crypto_risk_decisions ORDER BY created_at DESC"),
    ("Setup Events", "performance_setups"),
    ("Phase 2 Shadow Insights", "shadow_insights"),
    ("Phase 2 Shadow Portfolios", "shadow_portfolio_observations"),
    ("Phase 2 Sleeve Overlap", "shadow_overlap_observations"),
    ("Phase 2 Promotion Gates", "shadow_promotion_assessments"),
    ("Phase 3 Strategy States", "phase3_strategy_states"),
    ("Phase 3 Risk Decisions", "phase3_risk_decisions"),
    ("Phase 3 Allocations", "phase3_strategy_allocations"),
    ("Phase 3 Equity Watermark", "account_equity_watermarks"),
    ("Phase 4 Estimates", "phase4_strategy_estimates"),
    ("Phase 4 Allocations", "phase4_allocation_decisions"),
    ("Phase 4 Covariance", "phase4_covariance_snapshots"),
    ("Phase 4 Stress", "phase4_stress_results"),
    ("Phase 4 Strategy States", "phase4_strategy_states"),
    ("Strategy Registry Snapshots", "strategy_registry_snapshots"),
    ("Strategy Registry Decisions", "strategy_registry_decisions"),
    ("Winner Stop History", "position_stop_history"),
    ("Winner Pyramiding Milestones", "pyramiding_milestones"),
    ("Winner Add Risk Decisions", "add_risk_decisions"),
    ("Trend Management Decisions", "trend_management_decisions"),
    ("Rotation Groups", "rotation_groups"),
    ("Rotation Steps", "rotation_steps"),
    ("Rotation Contingent Entries", "rotation_contingent_entries"),
    ("Rotation Events", "rotation_events"),
    ("Rotation Group Approvals", "rotation_group_approvals"),
    ("Suppressed Setups", "SELECT * FROM performance_setups WHERE proposed=0 OR action_decision IN ('suppressed','blocked','shadow_only','failed_final_validation') ORDER BY created_at DESC"),
    ("Proposal Activity Status", "SELECT date('now') AS audit_date, (SELECT COUNT(*) FROM trade_proposals WHERE side='buy' AND date(created_at)=date('now')) AS buy_proposals_today, (SELECT COUNT(*) FROM trade_proposals WHERE side='buy' AND status='pending' AND (expires_at IS NULL OR datetime(expires_at)>datetime('now'))) AS active_pending_buy_proposals, (SELECT COUNT(*) FROM performance_setups WHERE date(created_at)=date('now')) AS setups_tracked_today, (SELECT COUNT(*) FROM performance_setups WHERE date(created_at)=date('now') AND proposed=0) AS suppressed_setups_today, (SELECT blocker FROM performance_blockers WHERE date(created_at)=date('now') GROUP BY blocker ORDER BY COUNT(*) DESC LIMIT 1) AS top_blocker_today, 'proposal_count_uncapped' AS proposal_policy"),
    ("Proposal Bottleneck Summary", "SELECT blocker, COUNT(*) AS setup_count, COUNT(DISTINCT symbol) AS symbol_count, MIN(created_at) AS first_seen, MAX(created_at) AS last_seen FROM performance_blockers GROUP BY blocker ORDER BY setup_count DESC"),
    ("Suppressed Setup Blockers", "SELECT ps.symbol, ps.tier, ps.setup_type, ps.action_decision, ps.score, ps.not_proposed_reason, pb.blocker, pb.reason, ps.created_at FROM performance_setups ps JOIN performance_blockers pb ON pb.setup_id=ps.id WHERE ps.proposed=0 OR ps.action_decision IN ('suppressed','blocked','shadow_only','failed_final_validation') ORDER BY ps.created_at DESC"),
    ("Risk or Workflow Blockers", "SELECT ps.symbol, ps.tier, ps.setup_type, ps.score, ps.not_proposed_reason, pb.blocker, pb.reason, ps.created_at FROM performance_setups ps JOIN performance_blockers pb ON pb.setup_id=ps.id WHERE pb.blocker IN ('risk_gate','cluster_gate','exposure_gate','max_daily_trades','cooldown') OR lower(pb.reason) LIKE '%risk%' OR lower(pb.reason) LIKE '%exposure%' OR lower(pb.reason) LIKE '%cooldown%' ORDER BY ps.created_at DESC"),
    ("Proposal Frequency Audit", "SELECT symbol, signal, score, classification, proposal_generated, no_action_reason, dedupe_status, dedupe_reason, cooldown_applied, cooldown_remaining_minutes, cooldown_reason, candidate_suppression_reason, final_proposal_message_category, created_at FROM market_memory ORDER BY created_at DESC"),
    ("Counterfactual Trades", "performance_counterfactuals"),
    ("Forward Returns 1D", "SELECT ps.symbol, ps.setup_type, ps.action_decision, fr.* FROM performance_forward_returns fr JOIN performance_setups ps ON ps.id=fr.setup_id WHERE fr.horizon_days=1 ORDER BY fr.due_at DESC"),
    ("Forward Returns 5D", "SELECT ps.symbol, ps.setup_type, ps.action_decision, fr.* FROM performance_forward_returns fr JOIN performance_setups ps ON ps.id=fr.setup_id WHERE fr.horizon_days=5 ORDER BY fr.due_at DESC"),
    ("Forward Returns 20D", "SELECT ps.symbol, ps.setup_type, ps.action_decision, fr.* FROM performance_forward_returns fr JOIN performance_setups ps ON ps.id=fr.setup_id WHERE fr.horizon_days=20 ORDER BY fr.due_at DESC"),
    ("MFE MAE Summary", "SELECT ps.symbol, ps.setup_type, ps.action_decision, fr.horizon_days, fr.max_favorable_excursion, fr.max_adverse_excursion, fr.hypothetical_stop_hit, fr.hypothetical_target_hit, fr.status FROM performance_forward_returns fr JOIN performance_setups ps ON ps.id=fr.setup_id ORDER BY ps.symbol, fr.horizon_days"),
    ("Proposal vs Suppressed Outcomes", "SELECT ps.symbol, ps.setup_type, ps.action_decision, ps.proposed, po.status, po.actual_or_shadow, ps.score, ps.not_proposed_reason FROM performance_setups ps LEFT JOIN performance_outcomes po ON po.setup_id=ps.id ORDER BY ps.created_at DESC"),
    ("Score Band Outcomes", "SELECT CASE WHEN score >= 90 THEN '90+' WHEN score >= 80 THEN '80-89' WHEN score >= 70 THEN '70-79' WHEN score >= 65 THEN '65-69' WHEN score >= 55 THEN '55-64' ELSE '<55' END AS score_band, action_decision, COUNT(*) AS setups, SUM(proposed) AS proposed FROM performance_setups GROUP BY score_band, action_decision ORDER BY score_band DESC, action_decision"),
    ("Blocker Outcome Analysis", "SELECT blocker, COUNT(*) AS setups, COUNT(DISTINCT symbol) AS symbols, MAX(created_at) AS latest_seen FROM performance_blockers GROUP BY blocker ORDER BY setups DESC"),
    ("Add-to-Winner Outcomes", "SELECT ps.*, po.status AS outcome_status FROM performance_setups ps LEFT JOIN performance_outcomes po ON po.setup_id=ps.id WHERE ps.setup_type='add_to_winner' ORDER BY ps.created_at DESC"),
    ("Exit Signal Outcomes", "SELECT ps.*, po.status AS outcome_status FROM performance_setups ps LEFT JOIN performance_outcomes po ON po.setup_id=ps.id WHERE ps.setup_type='exit' ORDER BY ps.created_at DESC"),
    ("Shadow Trades", "shadow_trades"),
    ("Actual vs Shadow", "SELECT * FROM trade_outcomes"),
    ("Forward Returns", "SELECT symbol, actual_or_shadow, entry_time, entry_price, forward_return_1d, forward_return_5d, forward_return_20d, outcome_status FROM trade_outcomes"),
    ("MAE MFE", "SELECT symbol, actual_or_shadow, entry_time, entry_price, max_favorable_excursion, max_adverse_excursion, stop_hit, target_reached FROM trade_outcomes"),
    ("Score Band Performance", "SELECT t.symbol, t.actual_or_shadow, COALESCE(s.score, json_extract(p.payload, '$.score')) as score, t.forward_return_1d, t.forward_return_5d, t.forward_return_20d, t.outcome_status FROM trade_outcomes t LEFT JOIN shadow_trades s ON t.trade_id = s.id LEFT JOIN orders o ON t.trade_id = o.id LEFT JOIN trade_proposals p ON o.proposal_id = p.id"),
    ("Symbol Performance", "SELECT symbol, actual_or_shadow, COUNT(*) as count, AVG(forward_return_20d) as avg_return_20d FROM trade_outcomes GROUP BY symbol, actual_or_shadow"),
    ("Add-on Opportunities", "add_on_opportunities"),
    ("Portfolio Exposure", "portfolio_exposure_snapshots"),
    ("Position Sizing Decisions", "position_sizing_decisions"),
    ("Candidate Ranking Decisions", "candidate_rankings"),
    ("Ranked Opportunity Sets", "ranked_opportunity_sets"),
    ("Profitability Decisions", "candidate_profitability_decisions"),
    ("Trade Economics", "trade_economics_records"),
    ("Strategy Trade Records", "strategy_trade_records"),
    ("Strategy Scorecards", "strategy_performance_snapshots"),
    ("Strategy Policies", "strategy_policy_decisions"),
    ("Validation Families", "profitability_validation_families"),
    ("Validation Decisions", "profitability_validation_decisions"),
    ("Validation Folds", "profitability_validation_folds"),
    ("Profit Attribution", "profit_attribution_records"),
    ("Proposal Batches", "proposal_batches"),
    ("Batch Candidates", "proposal_batch_candidates"),
    ("Risk Budget Decisions", "candidate_risk_budget_decisions"),
    ("Batch Approval Actions", "approval_batch_actions"),
    ("Candidate Allocation Decisions", "candidate_batch_allocations"),
    ("Risk Budget Snapshots", "risk_budget_snapshots"),
    ("Position Management State", "position_management_state"),
    ("Position Management Decisions", "position_management_decisions"),
    ("Profit Exit Events", "profit_exit_events"),
    ("Healthy Pullback Adds", "SELECT * FROM position_management_decisions WHERE decision_type='HEALTHY_PULLBACK_ADD'"),
    ("Profit Protection Events", "SELECT * FROM position_management_decisions WHERE decision_type='PROFIT_PROTECT_EXIT'"),
    ("Trailing Stop Events", "SELECT * FROM position_management_decisions WHERE decision_type='TRAILING_STOP_EXIT'"),
    ("Dynamic Universe Summary", "SELECT tier, COUNT(*) AS symbols, SUM(executable) AS executable_symbols, SUM(observation_only) AS observation_only_symbols, MAX(updated_at) AS latest_update FROM universe_symbols GROUP BY tier ORDER BY tier"),
    ("Dynamic Research Summary", "SELECT tier AS current_stage, universe_lane, COUNT(*) AS symbols, AVG(score) AS avg_score, MAX(updated_at) AS latest_update FROM universe_symbols GROUP BY tier, universe_lane ORDER BY tier, universe_lane"),
    ("Stage Semantics", "SELECT * FROM dynamic_universe_stage_semantics ORDER BY stage_order"),
    ("Universe Membership", "universe_membership_history"),
    ("Raw Universe Snapshot", "SELECT * FROM universe_symbols WHERE tier='raw_universe' ORDER BY score DESC, symbol"),
    ("Research Candidates", "SELECT * FROM universe_symbols WHERE tier='research_candidate' ORDER BY score DESC, symbol"),
    ("Research Candidate Briefs", "SELECT symbol,current_tier,universe_lane,research_score,rank,data_confidence,latest_price,price_freshness,dollar_volume,trend_summary,intraday_summary,relative_strength_vs_spy,sector,sector_relative_context,volatility_risk_summary,screener_reason,main_positive_reasons,main_blockers,missing_neutral_data,allowed_actions,blocked_actions,proposal_order_confirmation,next_expected_check,explanation_source,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Candidate Scores", "SELECT b.symbol,b.rank,b.research_score,b.data_confidence,s.liquidity_score,s.trend_score,s.intraday_momentum_score,s.relative_strength_score,s.volatility_quality_score,s.screener_mover_score,s.news_score,s.sector_theme_score,s.data_quality_score,b.created_at FROM research_candidate_briefs b LEFT JOIN symbol_research_scores s ON s.run_id=b.run_id AND s.symbol=b.symbol ORDER BY b.created_at DESC,b.rank"),
    ("Candidate Data Coverage", "SELECT symbol,data_confidence,price_freshness,liquidity_metrics,dollar_volume,trend_summary,intraday_summary,relative_strength_vs_spy,volatility_risk_summary,missing_neutral_data,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Candidate Endpoint Coverage", "SELECT symbol,endpoint_coverage,missing_neutral_data,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Candidate Promotion Requirements", "SELECT symbol,current_stage,next_stage_requirements,before_observation_requirements,before_paper_tradable_requirements,next_expected_check,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Candidate Block Reasons", "SELECT symbol,main_blockers,missing_neutral_data,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Candidate Next Steps", "SELECT symbol,current_stage,next_expected_check,allowed_actions,blocked_actions,proposal_order_confirmation,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Candidate Chart Data", "SELECT symbol,rank,research_score,data_confidence,dollar_volume,latest_price,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Research Funnel Chart Data", "SELECT 'raw_universe' AS stage, COUNT(*) AS symbols FROM universe_symbols WHERE tier='raw_universe' UNION ALL SELECT 'research_candidate', COUNT(*) FROM universe_symbols WHERE tier='research_candidate' UNION ALL SELECT 'observation', COUNT(*) FROM universe_symbols WHERE tier='observation' UNION ALL SELECT 'paper_tradable', COUNT(*) FROM universe_symbols WHERE tier='paper_tradable' UNION ALL SELECT 'trade_proposal', COUNT(*) FROM trade_proposals UNION ALL SELECT 'orders', COUNT(*) FROM orders"),
    ("Data Confidence Chart Data", "SELECT COALESCE(data_confidence,'unknown') AS data_confidence, COUNT(*) AS symbols FROM universe_symbols GROUP BY COALESCE(data_confidence,'unknown') ORDER BY symbols DESC"),
    ("Candidate Score Chart Data", "SELECT symbol,research_score,rank,data_confidence,created_at FROM research_candidate_briefs ORDER BY created_at DESC, rank"),
    ("Block Reason Chart Data", "SELECT main_blockers AS block_reason, COUNT(*) AS candidates FROM research_candidate_briefs GROUP BY main_blockers ORDER BY candidates DESC"),
    ("Observation Symbols", "SELECT * FROM universe_symbols WHERE tier='observation' ORDER BY score DESC, symbol"),
    ("Paper-Tradable Symbols", "SELECT * FROM universe_symbols WHERE tier='paper_tradable' ORDER BY score DESC, symbol"),
    ("Tier Summary", "SELECT tier, CASE WHEN tier='paper_tradable' THEN 'yes' ELSE 'no' END AS tradable, COUNT(*) AS symbols, SUM(CASE WHEN executable=1 THEN 1 ELSE 0 END) AS executable_symbols, SUM(CASE WHEN observation_only=1 THEN 1 ELSE 0 END) AS observation_only_symbols, AVG(score) AS avg_score, MAX(updated_at) AS latest_update FROM universe_symbols GROUP BY tier ORDER BY CASE tier WHEN 'paper_tradable' THEN 1 WHEN 'observation' THEN 2 WHEN 'research_candidate' THEN 3 WHEN 'global_research_only' THEN 4 ELSE 5 END"),
    ("Static Paper-Tradable Symbols", "SELECT symbol,tier,'yes' AS tradable,'static core' AS source_type,source,score,data_confidence,provider_health_status,last_successful_research_at,updated_at,'proposal requires setup, RiskEngine, Telegram approval, and final validation' AS proposal_status FROM universe_symbols WHERE tier='paper_tradable' AND source='existing_static_watchlist' ORDER BY symbol"),
    ("Dynamic Paper-Tradable Symbols", "SELECT symbol,tier,'yes' AS tradable,'dynamic promotion' AS source_type,source,score,data_confidence,last_promoted_at,provider_health_status,last_successful_research_at,updated_at,'proposal requires setup, RiskEngine, Telegram approval, and final validation' AS proposal_status FROM universe_symbols WHERE tier='paper_tradable' AND COALESCE(source,'')!='existing_static_watchlist' ORDER BY score DESC, symbol"),
    ("XL Sector Observation Audit", "SELECT r.symbol,r.current_tier,r.decision,r.reason,r.score,r.data_confidence,r.observation_since,r.observation_cycles,r.market_open_refreshes,r.eod_available,r.intraday_available,r.trend_summary,r.intraday_summary,r.liquidity_summary,r.volatility_summary,r.relative_strength_spy,r.relative_strength_qqq,r.cluster,r.cluster_exposure_blocker,r.promotion_requirements_met,r.promotion_requirements_missing,r.demotion_risk_reasons,r.tradable_status,r.proposal_allowed_status,r.proposal_block_reason,r.next_promotion_review_at,r.next_demotion_review_at,r.created_at FROM dynamic_universe_stage_reviews r INNER JOIN (SELECT symbol, MAX(created_at) AS created_at FROM dynamic_universe_stage_reviews WHERE symbol LIKE 'XL%' GROUP BY symbol) latest ON latest.symbol=r.symbol AND latest.created_at=r.created_at ORDER BY r.symbol"),
    ("Observation Maturity Review", "SELECT * FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Promotion Review Status", "SELECT symbol,current_tier,decision,reason,promotion_requirements_met,promotion_requirements_missing,last_promotion_review_at,next_promotion_review_at,tradable_status,proposal_allowed_status,proposal_block_reason,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Demotion Review Status", "SELECT symbol,current_tier,decision,reason,demotion_risk_reasons,demotion_guard_active,last_demotion_review_at,next_demotion_review_at,tradable_status,proposal_allowed_status,proposal_block_reason,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Observation Promotion Decisions", "SELECT symbol,from_tier,to_tier,score,reason,payload,created_at FROM symbol_promotion_decisions WHERE to_tier='observation' ORDER BY created_at DESC"),
    ("Observation Keep Reasons", "SELECT symbol,current_tier,decision,reason,next_stage_blocker,promotion_requirements_missing,created_at FROM dynamic_universe_stage_reviews WHERE decision LIKE 'keep_observation%' ORDER BY created_at DESC, symbol"),
    ("Observation Demotion Decisions", "SELECT symbol,current_tier,decision,reason,demotion_risk_reasons,created_at FROM dynamic_universe_stage_reviews WHERE decision LIKE '%demotion%' OR demotion_risk_reasons NOT IN ('[]','') ORDER BY created_at DESC, symbol"),
    ("Paper-Tradable Demotion Review", "SELECT symbol,current_tier,decision,reason,demotion_risk_reasons,demotion_guard_active,tradable_status,proposal_allowed_status,proposal_block_reason,created_at FROM dynamic_universe_stage_reviews WHERE current_tier='paper_tradable' ORDER BY created_at DESC, symbol"),
    ("Stage Decision History", "SELECT symbol,current_tier,review_type,decision,reason,score,data_confidence,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Promotion Block Reasons", "SELECT symbol,current_tier,promotion_requirements_missing,next_stage_blocker,reason,created_at FROM dynamic_universe_stage_reviews WHERE promotion_requirements_missing NOT IN ('[]','') ORDER BY created_at DESC, symbol"),
    ("Promotion Freshness Paths", "SELECT symbol,current_tier,decision,score,promotion_freshness_path,promotion_confidence_adjustment,promotion_data_limitations,eod_freshness,intraday_freshness,alpaca_quote_freshness,alpaca_tradability_result,proposal_allowed_status,proposal_block_reason,next_review_time,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Promotion Fallback Decisions", "SELECT symbol,current_tier,decision,reason,fallback_used,promotion_freshness_path,promotion_confidence_adjustment,promotion_data_limitations,next_review_time,created_at FROM dynamic_universe_stage_reviews WHERE COALESCE(fallback_used,'')='yes' OR promotion_freshness_path IN ('cached_intraday','alpaca_quote_fallback','eod_only_market_closed') ORDER BY created_at DESC, symbol"),
    ("Alpaca Quote Fallback Checks", "SELECT symbol,current_tier,decision,alpaca_quote_freshness,alpaca_tradability_result,promotion_freshness_path,proposal_block_reason,created_at FROM dynamic_universe_stage_reviews WHERE alpaca_quote_freshness IS NOT NULL ORDER BY created_at DESC, symbol"),
    ("Cached Intraday Usage", "SELECT symbol,current_tier,decision,intraday_freshness,promotion_freshness_path,promotion_confidence_adjustment,created_at FROM dynamic_universe_stage_reviews WHERE promotion_freshness_path='cached_intraday' OR intraday_freshness='cached' ORDER BY created_at DESC, symbol"),
    ("EOD-Only Market-Closed Reviews", "SELECT symbol,current_tier,decision,eod_freshness,promotion_freshness_path,promotion_confidence_adjustment,proposal_block_reason,created_at FROM dynamic_universe_stage_reviews WHERE promotion_freshness_path='eod_only_market_closed' ORDER BY created_at DESC, symbol"),
    ("Proposal Fresh Validation Blocks", "SELECT symbol,current_tier,decision,promotion_freshness_path,proposal_allowed_status,proposal_block_reason,proposal_block_reason_after_promotion,created_at FROM dynamic_universe_stage_reviews WHERE proposal_allowed_status!='allowed' ORDER BY created_at DESC, symbol"),
    ("Observation Promotion Block Reasons", "SELECT symbol,current_tier,decision,reason,promotion_requirements_missing,next_stage_blocker,next_review_time,created_at FROM dynamic_universe_stage_reviews WHERE current_tier='observation' AND decision LIKE 'keep_observation%' ORDER BY created_at DESC, symbol"),
    ("Global Research-Only Updates", "SELECT symbol,tier,universe_lane,alpaca_compatible,executable,observation_only,score,source,exclusion_reason,updated_at FROM universe_symbols WHERE universe_lane='global_research_only' ORDER BY updated_at DESC, symbol"),
    ("Static Reconciliation Events", "SELECT symbol,from_tier,to_tier,score,reason,created_at FROM symbol_promotion_decisions WHERE json_extract(payload,'$.existing_static')=1 ORDER BY created_at DESC"),
    ("Dynamic Promotion Events", "SELECT symbol,from_tier,to_tier,score,reason,json_extract(payload,'$.promotion_freshness_path') AS promotion_freshness_path,json_extract(payload,'$.fallback_used') AS fallback_used,json_extract(payload,'$.proposal_block_reason_after_promotion') AS proposal_block_reason,created_at FROM symbol_promotion_decisions WHERE COALESCE(json_extract(payload,'$.existing_static'),0)!=1 ORDER BY created_at DESC"),
    ("Demotion Risk Reasons", "SELECT symbol,current_tier,demotion_risk_reasons,demotion_guard_active,reason,created_at FROM dynamic_universe_stage_reviews WHERE demotion_risk_reasons NOT IN ('[]','') OR demotion_guard_active=1 ORDER BY created_at DESC, symbol"),
    ("Tradability Status", "SELECT symbol,current_tier,tradable_status,proposal_allowed_status,proposal_block_reason,decision,reason,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Proposal Eligibility Status", "SELECT symbol,current_tier,proposal_allowed_status,proposal_block_reason,tradable_status,decision,reason,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Digest Tier Snapshot", "SELECT symbol,tier,CASE WHEN tier='paper_tradable' THEN 'yes' ELSE 'no' END AS tradable,CASE WHEN tier='paper_tradable' THEN 'blocked until setup/risk/approval/final validation pass' WHEN tier='observation' THEN 'no: needs paper-tradable promotion' WHEN tier='research_candidate' THEN 'no: needs observation promotion first' ELSE 'no' END AS proposal_status,source,universe_lane,score,data_confidence,provider_health_status,updated_at FROM universe_symbols WHERE tier IN ('paper_tradable','observation','research_candidate') ORDER BY CASE tier WHEN 'paper_tradable' THEN 1 WHEN 'observation' THEN 2 ELSE 3 END, score DESC, symbol"),
    ("Provider Health Deduped", "SELECT provider,status,COUNT(*) AS events,MAX(checked_at) AS latest_checked,MAX(error) AS latest_error FROM data_provider_health GROUP BY provider,status ORDER BY latest_checked DESC"),
    ("Universe Events Timeline", "SELECT created_at,'promotion' AS event_type,symbol,from_tier AS from_state,to_tier AS to_state,reason FROM symbol_promotion_decisions UNION ALL SELECT created_at,'demotion' AS event_type,symbol,from_tier AS from_state,to_tier AS to_state,reason FROM symbol_demotion_decisions UNION ALL SELECT created_at,review_type AS event_type,symbol,current_tier AS from_state,decision AS to_state,reason FROM dynamic_universe_stage_reviews ORDER BY created_at DESC"),
    ("EODHD Historical Metrics", "SELECT symbol,eod_available,latest_price,price_freshness,trend_summary,liquidity_summary,volatility_summary,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Relative Strength Metrics", "SELECT symbol,relative_strength_spy,relative_strength_qqq,score,data_confidence,created_at FROM dynamic_universe_stage_reviews ORDER BY created_at DESC, symbol"),
    ("Cluster Exposure Blockers", "SELECT symbol,current_tier,cluster,cluster_exposure_blocker,promotion_requirements_missing,reason,created_at FROM dynamic_universe_stage_reviews WHERE COALESCE(cluster,'')!='' OR COALESCE(cluster_exposure_blocker,'')!='' ORDER BY created_at DESC, symbol"),
    ("Demoted Symbols", "SELECT * FROM universe_symbols WHERE tier='demoted' ORDER BY updated_at DESC, symbol"),
    ("Symbol Research Scores", "symbol_research_scores"),
    ("News Events", "symbol_news_events"),
    ("Trend Snapshots", "symbol_trend_snapshots"),
    ("Sector Regime", "sector_regime_snapshots"),
    ("Promotion Decisions", "symbol_promotion_decisions"),
    ("Demotion Decisions", "symbol_demotion_decisions"),
    ("Dynamic Universe Audit", "dynamic_universe_audit"),
    ("Data Provider Health", "data_provider_health"),
    ("Provider Capabilities", "data_provider_capabilities"),
    ("Endpoint Availability", "SELECT provider,endpoint_name,available,plan_limited,last_success_at,last_failure_at,failure_count,last_status_code,last_error_category,disabled_until,used_for_scoring,updated_at FROM data_provider_capabilities ORDER BY provider, endpoint_name"),
    ("Dynamic Universe Performance", "dynamic_universe_performance"),
    ("Dynamic Universe Schedule State", "dynamic_universe_schedule_state"),
    ("Latest Dynamic Universe Subtask Status", "SELECT schedule_name,schedule_type,last_started_at,last_completed_at,last_success_at,last_skipped_at,last_skip_reason,CASE WHEN last_skip_reason IS NOT NULL AND (last_success_at IS NULL OR datetime(last_skipped_at)>datetime(last_success_at)) THEN 'current_skip' WHEN last_success_at IS NOT NULL THEN 'latest_success' ELSE 'unknown' END AS current_status,data_freshness_status,provider_health_status,promotion_allowed,demotion_allowed,updated_at FROM dynamic_universe_schedule_state ORDER BY schedule_name"),
    ("Research Subtask Skip Reasons", "SELECT event_type,json_extract(detail,'$.run_type') AS run_type,json_extract(detail,'$.reason') AS reason,created_at,CASE WHEN EXISTS (SELECT 1 FROM universe_research_runs r WHERE r.research_type=json_extract(dynamic_universe_audit.detail,'$.run_type') AND r.status='completed' AND datetime(r.ended_at)>datetime(dynamic_universe_audit.created_at)) THEN 'historical_superseded' ELSE 'current_or_unrecovered' END AS skip_status FROM dynamic_universe_audit WHERE event_type IN ('dynamic_universe_research_skipped','dynamic_universe_research_missed') ORDER BY created_at DESC"),
    ("Stale Research Guard Status", "SELECT event_type,detail,created_at,'Blocks BUY/ADD eligibility and unsafe paper-tradable promotion; observation-only tracking and SELL/EXIT monitoring may continue' AS guard_semantics FROM dynamic_universe_audit WHERE event_type IN ('dynamic_universe_stale_data_guard','dynamic_universe_promotions_blocked_stale_research') ORDER BY created_at DESC"),
    ("Provider State Recovery", "SELECT event_type,detail,created_at,'dynamic_universe_audit' AS source_table FROM dynamic_universe_audit WHERE event_type='provider_missing_key_state_recovered' UNION ALL SELECT event_type,detail,created_at,'audit_events' AS source_table FROM audit_events WHERE event_type='provider_missing_key_state_recovered' ORDER BY created_at DESC"),
    ("Observation Promotion Source", "SELECT p.symbol,p.from_tier,p.to_tier,p.score,p.reason,json_extract(p.payload,'$.source') AS source,json_extract(p.payload,'$.data_confidence') AS data_confidence,json_extract(p.payload,'$.endpoint_coverage') AS endpoint_coverage,p.created_at FROM symbol_promotion_decisions p WHERE p.to_tier='observation' ORDER BY p.created_at DESC"),
    ("Candidate Promotion Trace", "SELECT p.symbol,p.from_tier,p.to_tier,p.score,p.reason,json_extract(p.payload,'$.source') AS source,json_extract(p.payload,'$.universe_lane') AS universe_lane,json_extract(p.payload,'$.endpoint_coverage') AS endpoint_coverage,p.created_at FROM symbol_promotion_decisions p ORDER BY p.created_at DESC"),
    ("Digest Status Semantics", "SELECT 'full_research_skipped' AS label,'Use only when no Dynamic Universe research subtask completed and no promotions were recorded in the digest window' AS meaning UNION ALL SELECT 'subtask_skipped','Use when one schedule subtask skipped but existing/completed research state was still available' UNION ALL SELECT 'existing_state_used','Use when observation promotions rely on deterministic candidate state from a completed scan or refresh' UNION ALL SELECT 'stale_research_guard','Blocks BUY/ADD eligibility and unsafe paper-tradable promotion; does not block observation-only tracking or SELL/EXIT monitoring'"),
    ("Missed Research Cycles", "SELECT * FROM dynamic_universe_schedule_state WHERE missed_count > 0 OR catchup_required=1 ORDER BY updated_at DESC"),
    ("Catch-Up Runs", "SELECT * FROM dynamic_universe_schedule_state WHERE catchup_attempted_at IS NOT NULL OR catchup_completed_at IS NOT NULL ORDER BY updated_at DESC"),
    ("Stale Research Guards", "SELECT * FROM dynamic_universe_audit WHERE event_type='dynamic_universe_stale_data_guard' ORDER BY created_at DESC"),
    ("Dynamic Universe Promotion Blocks", "SELECT * FROM dynamic_universe_audit WHERE event_type='dynamic_universe_promotions_blocked_stale_research' ORDER BY created_at DESC"),
    ("Dynamic Universe Demotion Blocks", "SELECT * FROM dynamic_universe_audit WHERE event_type='dynamic_universe_demotions_blocked_provider_unavailable' ORDER BY created_at DESC"),
    ("Research Candidate Blocks", "research_candidate_block_reasons"),
    ("Data Confidence", "SELECT symbol,tier,universe_lane,score,data_confidence,data_confidence_reason,data_freshness_status,provider_health_status,promotion_allowed,demotion_allowed,updated_at FROM universe_symbols ORDER BY updated_at DESC, score DESC"),
    ("Top Near-Miss Symbols", "SELECT * FROM dynamic_universe_audit WHERE event_type='dynamic_universe_near_miss_symbols' ORDER BY created_at DESC"),
    ("Dynamic Universe Source Coverage", "SELECT source,tier,universe_lane,data_confidence,COUNT(*) AS symbols,AVG(score) AS avg_score,MAX(updated_at) AS latest_update FROM universe_symbols GROUP BY source,tier,universe_lane,data_confidence ORDER BY source,tier,universe_lane,data_confidence"),
    ("Symbol Intake Classification", "SELECT symbol, provider_symbol, exchange, asset_class, region, currency, universe_lane, alpaca_compatible, exclusion_reason, tier, source, score, updated_at FROM universe_symbols ORDER BY universe_lane, score DESC, symbol"),
    ("Alpaca-Compatible Candidates", "SELECT * FROM universe_symbols WHERE universe_lane='alpaca_compatible_us' ORDER BY score DESC, symbol"),
    ("Global Research-Only Symbols", "SELECT * FROM universe_symbols WHERE universe_lane='global_research_only' ORDER BY score DESC, symbol"),
    ("Excluded Symbols", "SELECT * FROM universe_symbols WHERE universe_lane='excluded_or_low_quality' ORDER BY updated_at DESC, symbol"),
    ("Symbol Exclusion Reasons", "SELECT exclusion_reason, COUNT(*) AS symbols, MAX(updated_at) AS latest_update FROM universe_symbols WHERE universe_lane='excluded_or_low_quality' GROUP BY exclusion_reason ORDER BY symbols DESC"),
    ("Near-Miss US Candidates", "SELECT symbol, score, data_confidence, block_reason, liquidity_score, trend_score, intraday_momentum_score, relative_strength_score, volatility_quality_score, screener_mover_score, news_score, created_at FROM research_candidate_block_reasons WHERE universe_lane='alpaca_compatible_us' ORDER BY score DESC, created_at DESC LIMIT 50"),
    ("Near-Miss Global Research", "SELECT symbol, score, data_confidence, block_reason, created_at FROM research_candidate_block_reasons WHERE universe_lane='global_research_only' ORDER BY score DESC, created_at DESC LIMIT 50"),
    ("Candidate Block Reason Summary", "SELECT COALESCE(universe_lane,'unknown') AS universe_lane, block_reason, COUNT(*) AS symbols, AVG(score) AS avg_score, MAX(created_at) AS latest_seen FROM research_candidate_block_reasons GROUP BY COALESCE(universe_lane,'unknown'), block_reason ORDER BY symbols DESC"),
    ("Research Rule Strictness Audit", "SELECT block_reason, COUNT(*) AS blocked_symbols, CASE WHEN block_reason IN ('missing or stale price data','missing liquidity data','liquidity below minimum','price below minimum') THEN 'hard safety/data gate' WHEN block_reason LIKE '%excluded%' OR block_reason LIKE '%otc%' THEN 'hard intake gate' ELSE 'soft discovery threshold' END AS recommended_gate_type, MAX(created_at) AS latest_seen FROM research_candidate_block_reasons GROUP BY block_reason ORDER BY blocked_symbols DESC"),
    ("Exploration Candidates", "SELECT * FROM universe_symbols WHERE tier='research_candidate' AND COALESCE(source,'') NOT LIKE 'existing_static%' ORDER BY score DESC, symbol"),
    ("Provider Capability Usage", "SELECT provider,endpoint_name,available,plan_limited,last_error_category,disabled_until,used_for_scoring,updated_at FROM data_provider_capabilities ORDER BY provider, endpoint_name"),
    ("Optional News Provider Status", "SELECT 'marketaux' AS provider, 'disabled_until_key_exists' AS status, 'TradingAgent.MARKETAUX_API_KEY' AS key_source, 'news fallback only for shortlisted symbols' AS usage"),
]


SECRET_VALUE_PATTERNS = (
    re.compile(r"sk-[A-Za-z0-9_-]{20,}"),
    re.compile(r"[0-9]{6,}:[A-Za-z0-9_-]{10,}"),
    re.compile(r"PK[A-Z0-9]{15,}"),
)

TELEGRAM_TEXT_KEYS = {"raw_message", "raw_command", "raw_command_redacted", "text"}
TELEGRAM_ID_KEYS = {"sender_id", "updated_by", "telegram_user_id", "chat_id", "from_id"}
SENSITIVE_KEY_PARTS = ("key", "secret", "token", "password", "account_id")
TEXT_BLOB_KEYS = {
    "telegram_message",
    "message_text",
    "proposal_text",
    "formatted_message",
    "raw_message",
    "raw_text",
    "text",
    "review_text",
    "ai_review_text",
    "gpt_response_text",
    "approval_text",
    "command_text",
    "summary",
    "reason",
    "reasoning_notes",
    "main_risk",
    "risks",
    "warnings",
    "messages",
}
JSON_LIKE_HEADERS = {
    "payload",
    "raw_payload",
    "request_payload",
    "response_payload",
    "telegram_payload",
    "values_json",
    "config_json",
    "detail",
    "risks",
    "portfolio_state_json",
    "single_symbol_exposure_json",
    "cluster_exposure_json",
    "score_components",
    "signal_state",
    "trend_metrics",
    "volatility_metrics",
    "liquidity_metrics",
    "relative_strength_metrics",
    "portfolio_exposure",
    "cluster_exposure",
    "risk_budget",
}


def _looks_like_report_text_blob(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    if len(stripped) >= 32 and " " in stripped:
        return True
    words = stripped.split()
    if len(words) >= 4:
        return True
    if any(mark in stripped for mark in (". ", "! ", "? ", "\n")):
        return True
    return False


def redact_report_payload(value: Any) -> Any:
    if isinstance(value, dict):
        redacted: dict[str, Any] = {}
        for key, item in value.items():
            key_lower = str(key).lower()
            if key_lower in TELEGRAM_TEXT_KEYS:
                redacted[f"[REDACTED TELEGRAM TEXT FIELD:{len(redacted)}]"] = "[REDACTED TELEGRAM TEXT]"
            elif key_lower in TEXT_BLOB_KEYS or "payload" in key_lower:
                if isinstance(item, (dict, list)):
                    redacted[key] = redact_report_payload(item)
                else:
                    redacted[key] = "[REDACTED TEXT]"
            elif key_lower in TELEGRAM_ID_KEYS or key_lower.endswith("_sender_id"):
                redacted[f"[REDACTED ID FIELD:{len(redacted)}]"] = "[REDACTED ID]"
            elif any(part in key_lower for part in SENSITIVE_KEY_PARTS):
                redacted[key] = "[REDACTED]"
            else:
                redacted[key] = redact_report_payload(item)
        return redacted
    if isinstance(value, list):
        redacted_items: list[Any] = []
        for item in value:
            if isinstance(item, str) and _looks_like_report_text_blob(item):
                redacted_items.append("[REDACTED TEXT]")
            else:
                redacted_items.append(redact_report_payload(item))
        return redacted_items
    return redact(value)


def redact_report_value(table: str, header: str, value: Any, include_raw_telegram: bool = False) -> Any:
    header_lower = str(header).lower()
    if header_lower in TELEGRAM_TEXT_KEYS and not include_raw_telegram:
        return "[REDACTED TELEGRAM TEXT]"
    if header_lower in TEXT_BLOB_KEYS and not include_raw_telegram:
        return "[REDACTED TEXT]"
    if header_lower in TELEGRAM_ID_KEYS:
        return "[REDACTED ID]"
    if value is None:
        return None
    if header_lower in JSON_LIKE_HEADERS and isinstance(value, str):
        try:
            value = json_dumps(redact_report_payload(json.loads(value)))
        except (ValueError, TypeError):
            value = "[REDACTED TEXT]" if ("payload" in header_lower or header_lower in TEXT_BLOB_KEYS) else redact(value)
    if isinstance(value, str):
        for pattern in SECRET_VALUE_PATTERNS:
            value = pattern.sub("[REDACTED SECRET]", value)
    return value


def _write_rows(sheet: Any, rows: list[dict[str, Any]], table: str, include_raw_telegram: bool = False) -> None:
    if not rows:
        sheet.append(["No records"])
        return
    headers = list(rows[0])
    display_headers = []
    for header in headers:
        header_lower = str(header).lower()
        if header_lower in TELEGRAM_TEXT_KEYS and not include_raw_telegram:
            display_headers.append("[REDACTED TELEGRAM TEXT FIELD]")
        elif header_lower in TELEGRAM_ID_KEYS or header_lower.endswith("_sender_id"):
            display_headers.append("[REDACTED ID FIELD]")
        else:
            display_headers.append(header)
    sheet.append(display_headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        values = []
        for header in headers:
            value = row.get(header)
            if (
                table == "control_state"
                and header == "value"
                and str(row.get("key", "")).lower().endswith("last_command")
                and not include_raw_telegram
            ):
                value = "[REDACTED TELEGRAM TEXT]"
            else:
                value = redact_report_value(table, header, value, include_raw_telegram)
            values.append(value)
        sheet.append(values)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = min(50, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[column[0].column_letter].width = width


def export_excel(storage: Storage, config: dict[str, Any], output_path: str | Path | None = None, include_raw_telegram: bool = False) -> Path:
    output = Path(output_path or PROJECT_ROOT / "data" / "exports" / f"trading_agent_report_{date.today().isoformat()}.xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Summary Dashboard")
    orders = storage.fetch_all("SELECT * FROM orders")
    fills = storage.fetch_all("SELECT * FROM fills")
    proposals = storage.fetch_all("SELECT * FROM trade_proposals")
    positions = storage.fetch_all("SELECT * FROM positions ORDER BY created_at DESC")
    runs = storage.fetch_all("SELECT * FROM runs ORDER BY started_at DESC LIMIT 1")
    errors = storage.fetch_all("SELECT * FROM errors ORDER BY created_at DESC LIMIT 1")
    profiles = config.get("market_profiles", {})
    active_profile_key = "default"
    active_profile = {}
    for k, v in profiles.items():
        if v.get("status") == "active":
            active_profile_key = k
            active_profile = v
            break

    metrics = [
        ("Mode", config.get("mode", "paper")), ("Report date", date.today().isoformat()),
        ("Starting capital", "See broker statement"), ("Current equity", "See latest cash snapshot"),
        ("Realized PnL", "See Daily PnL"), ("Unrealized PnL", "See Positions"), ("Total PnL", "See Daily PnL"),
        ("Win rate", "Requires completed fills"), ("Number of proposals", len(proposals)),
        ("Approved trades", sum(p["status"] == "approved" for p in proposals)),
        ("Rejected trades", sum(p["status"] == "rejected" for p in proposals)),
        ("Expired proposals", sum(p["status"] == "expired" for p in proposals)), ("Orders placed", len(orders)),
        ("Fills", len(fills)), ("Max drawdown", "See Daily PnL"), ("Average gain", "Requires completed fills"),
        ("Average loss", "Requires completed fills"), ("Profit factor", "Requires completed fills"),
        ("Current open positions", len(positions)), ("Daily loss limit status", "See Risk Checks"),
        ("Weekly loss limit status", "See Risk Checks"), ("Last bot run", runs[0]["started_at"] if runs else "Never"),
        ("Last error", errors[0]["message"] if errors else "None"), ("System health status", "Review latest preflight checks"),
        ("Suggested cash-out amount", "Recommendation only; see Cash Management"),
        ("Active Market Profile", active_profile_key),
        ("Execution Enabled", "yes" if active_profile.get("execution_enabled") else "no"),
        ("Default Proposal Expiry (minutes)", config.get("proposal_expiry_default_minutes", 15)),
        ("Auto-execution Mode Status", config.get("auto_execution_mode", "manual_only")),
        ("Auto-execution Enabled", str(config.get("auto_execution_enabled", False))),
        ("Asset Universe Status (Tradeable)", ", ".join(active_profile.get("watchlist", [])) if active_profile else ""),
        ("Asset Universe Status (Observation)", ", ".join(active_profile.get("observation_watchlist", [])) if active_profile else ""),
    ]
    summary.append(["TradingAgent Summary Dashboard"])
    summary["A1"].font = Font(bold=True, size=16)
    summary.append(["Metric", "Value"])
    for cell in summary[2]: cell.font = Font(bold=True)
    for metric in metrics: summary.append(metric)
    summary.freeze_panes = "A3"
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 48
    for name, query_or_table in SHEETS[1:]:
        sheet = workbook.create_sheet(name)
        if query_or_table.startswith("SELECT"):
            _write_rows(sheet, storage.fetch_all(query_or_table), name.lower().replace(" ", "_"), include_raw_telegram)
        else:
            _write_rows(sheet, storage.fetch_all(f'SELECT * FROM "{query_or_table}"'), str(query_or_table), include_raw_telegram)
        sheet.conditional_formatting.add("A2:Z10000", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")))
    workbook.save(output)
    return output


def main() -> None:
    from .utils import load_config
    config = load_config()
    storage = Storage(PROJECT_ROOT / config["storage"]["sqlite_path"])
    storage.initialize()
    print(export_excel(storage, config))


if __name__ == "__main__":
    main()
