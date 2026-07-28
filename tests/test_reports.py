import json

import pytest
from openpyxl import load_workbook

import app.reports as reports
from app.reports import (
    EXCEL_INVALID_SHEET_TITLE,
    EXCEL_MAX_SHEET_TITLE_LENGTH,
    SHEETS,
    export_excel,
    redact_report_value,
)
from app.storage import Storage


def test_export_has_expected_sheets(tmp_path):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    path = export_excel(storage, {"mode": "paper"}, tmp_path / "report.xlsx")
    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == [name for name, _ in SHEETS]


def test_registered_report_sheet_titles_are_excel_compatible():
    titles = [name for name, _query in SHEETS]

    assert all(titles)
    assert all(len(title) <= EXCEL_MAX_SHEET_TITLE_LENGTH for title in titles)
    assert all(EXCEL_INVALID_SHEET_TITLE.search(title) is None for title in titles)
    assert all(
        not title.startswith("'") and not title.endswith("'") for title in titles
    )
    assert len({title.casefold() for title in titles}) == len(titles)


@pytest.mark.parametrize(
    "invalid_title",
    [
        "X" * (EXCEL_MAX_SHEET_TITLE_LENGTH + 1),
        "Invalid/Title",
        "'Apostrophe",
        "apostrophe'",
    ],
)
def test_export_rejects_invalid_sheet_title_before_writing(
    tmp_path, monkeypatch, invalid_title
):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    monkeypatch.setattr(
        reports,
        "SHEETS",
        [("Summary Dashboard", None), (invalid_title, "orders")],
    )

    with pytest.raises(ValueError, match="worksheet title"):
        export_excel(storage, {"mode": "paper"}, tmp_path / "report.xlsx")

    assert not (tmp_path / "report.xlsx").exists()


def test_export_rejects_case_insensitive_duplicate_sheet_title(
    tmp_path, monkeypatch
):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    monkeypatch.setattr(
        reports,
        "SHEETS",
        [("Summary Dashboard", None), ("summary dashboard", "orders")],
    )

    with pytest.raises(ValueError, match="case-insensitively unique"):
        export_excel(storage, {"mode": "paper"}, tmp_path / "report.xlsx")

    assert not (tmp_path / "report.xlsx").exists()


def test_report_redacts_telegram_text_and_sender_ids_inside_json_detail():
    detail = '{"raw_command":"sleep","sender_id":"authorized_user_123","nested":{"text":"awake","updated_by":"7"}}'

    redacted = redact_report_value("audit_events", "detail", detail)

    assert "sleep" not in redacted
    assert "awake" not in redacted
    assert "authorized_user_123" not in redacted
    assert "raw_command" not in redacted
    assert "sender_id" not in redacted
    assert "[REDACTED TELEGRAM TEXT FIELD" in redacted
    assert "[REDACTED ID FIELD" in redacted
    assert "updated_by" not in redacted


def test_export_redacts_control_state_last_command_value(tmp_path):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    storage.execute(
        "INSERT INTO control_state(key, value, updated_by, source, raw_command_redacted) VALUES(?,?,?,?,?)",
        ("sleep_mode_last_command", "awake", "authorized_user_123", "telegram", "awake"),
    )

    path = export_excel(storage, {"mode": "paper"}, tmp_path / "report.xlsx")
    workbook = load_workbook(path, read_only=True)
    rows = list(workbook["Control State"].iter_rows(values_only=True))

    assert rows[1][1] == "[REDACTED TELEGRAM TEXT]"
    assert rows[1][3] == "[REDACTED ID]"
    assert rows[1][5] == "[REDACTED TELEGRAM TEXT]"


def test_report_redaction_removes_embedded_text_but_preserves_numeric_fields():
    payload = {
        "telegram_message": "Buy SPY now",
        "proposal_text": "Generated proposal text",
        "review_text": "GPT narrative",
        "risks": ["This is a generated review sentence that should not export."],
        "score": 91,
        "suggested_notional": 15.0,
        "nested": {"message_text": "hello telegram", "risk_score": 12.5},
    }

    redacted = redact_report_value("trade_proposals", "payload", json.dumps(payload))

    assert "Buy SPY now" not in redacted
    assert "Generated proposal text" not in redacted
    assert "GPT narrative" not in redacted
    assert "generated review sentence" not in redacted
    assert '"score":91' in redacted
    assert '"suggested_notional":15.0' in redacted
    assert '"risk_score":12.5' in redacted


def test_export_redacts_nested_payload_text_in_workbook(tmp_path):
    storage = Storage(tmp_path / "payloads.db")
    storage.initialize()
    storage.execute(
        "INSERT INTO trade_proposals(id,run_id,symbol,side,notional,status,created_at,expires_at,strategy_version,payload) VALUES(?,?,?,?,?,?,?,?,?,?)",
        ("p1", "run-1", "SPY", "buy", 5.0, "pending", "2026-06-25T00:00:00+00:00", "2026-06-25T00:15:00+00:00", "rule_based_v1", json.dumps({"proposal_text": "Full proposal body", "score": 88, "telegram_message": "YES SPY"})),
    )
    storage.execute(
        "INSERT INTO ai_reviews(run_id,proposal_id,summary,risks,caution_level,payload,created_at) VALUES(?,?,?,?,?,?,?)",
        ("run-1", "p1", "This is a long GPT summary", json.dumps(["Choppy price action"]), "medium", json.dumps({"review_text": "Full GPT review", "main_risk": "Narrative risk", "score": 88}), "2026-06-25T00:00:01+00:00"),
    )

    path = export_excel(storage, {"mode": "paper"}, tmp_path / "report.xlsx")
    workbook = load_workbook(path, read_only=True)
    cell_text = "\n".join(
        str(cell)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )

    assert "Full proposal body" not in cell_text
    assert "YES SPY" not in cell_text
    assert "This is a long GPT summary" not in cell_text
    assert "Full GPT review" not in cell_text
    assert "Narrative risk" not in cell_text
    assert "88" in cell_text
