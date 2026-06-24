from __future__ import annotations

from datetime import UTC, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

import app.run_lock as run_lock
import app.service as service_module
from app.broker_alpaca import AlpacaBroker
from app.reconciliation import BrokerReconciler
from app.reports import export_excel
from app.risk_engine import RiskEngine
from app.service import TradingService
from app.storage import Storage
from app.strategy_rule_based import Signal
from app.telegram_bot import redact_telegram_update
from app.utils import load_config


class HealthyTelegram:
    def is_available(self, force=False):
        return True


class AuthoritativeBroker:
    def __init__(self, *, weekly_loss=2.0, margin=False):
        self.weekly_loss = weekly_loss
        cash = -1 if margin else 100
        long_value = 101 if margin else 0
        self.account = SimpleNamespace(
            buying_power="100", equity="100", last_equity="101", cash=str(cash),
            long_market_value=str(long_value), short_market_value="0",
        )

    def get_account(self): return self.account
    def get_positions(self): return []
    def get_open_orders(self): return []
    def get_clock(self): return SimpleNamespace(is_open=True)
    def is_market_open(self): return True
    def get_loss_metrics(self): return {"daily_loss": 1.0, "weekly_loss": self.weekly_loss}


def make_service(tmp_path, broker):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    service = TradingService.__new__(TradingService)
    service.config = load_config()
    service.storage = storage
    service.broker = broker
    service.run_id = "run"
    service.telegram = HealthyTelegram()
    service._context_cache = None
    service._auto_block_audited = False
    return service, storage


def test_final_context_uses_authoritative_values(tmp_path, monkeypatch):
    service, _ = make_service(tmp_path, AuthoritativeBroker(margin=True))
    monkeypatch.setattr(service_module, "internet_available", lambda: False)
    context = service._portfolio_context({"symbol": "SPY", "action": "entry"}, approval_valid=True)
    assert context["daily_loss"] == 1.0
    assert context["weekly_loss"] == 2.0
    assert context["uses_margin"] is True
    assert context["internet_available"] is False
    assert context["broker_available"] is True
    assert context["telegram_available"] is True
    assert context["database_writable"] is True


def test_unknown_weekly_loss_blocks_risk(tmp_path, monkeypatch, proposal, context):
    service, _ = make_service(tmp_path, AuthoritativeBroker(weekly_loss=None))
    monkeypatch.setattr(service_module, "internet_available", lambda: True)
    authoritative = service._portfolio_context({"symbol": "SPY", "action": "entry"}, approval_valid=True)
    context.update(authoritative)
    decision = RiskEngine(service.config).evaluate(proposal, context)
    assert not decision.passed
    assert any(check.name == "weekly_loss_known" and not check.passed for check in decision.checks)


class ReconcileBroker:
    def __init__(self, fail=False):
        self.fail = fail
        self.submit_calls = 0
        self.remote = SimpleNamespace(
            id="broker-1", status="filled", filled_qty="2", filled_avg_price="10.5",
            filled_at=datetime.now(UTC),
        )

    def get_order(self, order_id):
        if self.fail: raise RuntimeError("unknown")
        return self.remote

    def get_order_by_client_order_id(self, client_order_id): return self.get_order(client_order_id)
    def get_account(self): return SimpleNamespace(equity="100", cash="90")
    def get_positions(self): return []
    def submit_order(self, *args, **kwargs):
        self.submit_calls += 1
        raise AssertionError("reconciliation must never submit")


def insert_local_order(storage):
    storage.execute(
        "INSERT INTO orders(id,run_id,broker_order_id,client_order_id,symbol,side,notional,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
        ("local-1", "run", "broker-1", "client-1", "SPY", "buy", 5, "pending_new", datetime.now(UTC).isoformat(), datetime.now(UTC).isoformat()),
    )


def test_reconciliation_updates_order_and_inserts_fill_once(tmp_path):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    insert_local_order(storage)
    broker = ReconcileBroker()
    reconciler = BrokerReconciler(broker, storage, "run")
    reconciler.reconcile()
    reconciler.reconcile()
    assert storage.fetch_all("SELECT status FROM orders WHERE id='local-1'")[0]["status"] == "filled"
    assert storage.fetch_all("SELECT count(*) n FROM fills")[0]["n"] == 1
    assert broker.submit_calls == 0


def test_filled_batch_order_creates_actual_performance_outcome_and_links_rows(tmp_path):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    now = datetime.now(UTC).isoformat()
    storage.execute(
        "INSERT INTO trade_proposals(id,run_id,symbol,side,notional,status,created_at,expires_at,payload) VALUES(?,?,?,?,?,?,?,?,?)",
        (
            "proposal-1", "scan-run", "IWM", "buy", 15, "approved", now,
            (datetime.now(UTC) + timedelta(minutes=5)).isoformat(),
            '{"score":90,"asset_score":90,"latest_price":299.08,"reason":"ranked batch candidate"}',
        ),
    )
    storage.execute(
        "INSERT INTO proposal_batches(id,run_id,status,created_at,expires_at) VALUES(?,?,?,?,?)",
        ("batch-1", "scan-run", "completed", now, (datetime.now(UTC) + timedelta(minutes=5)).isoformat()),
    )
    storage.execute(
        "INSERT INTO proposal_batch_candidates(id,batch_id,proposal_id,candidate_symbol,candidate_side,candidate_action,candidate_status,rank,created_at,expires_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
        ("candidate-1", "batch-1", "proposal-1", "IWM", "buy", "entry", "submitted", 1, now, (datetime.now(UTC) + timedelta(minutes=5)).isoformat()),
    )
    storage.execute(
        "INSERT INTO approval_batch_actions(id,run_id,batch_id,proposal_id,action,status,created_at) VALUES(?,?,?,?,?,?,?)",
        ("batch-action-1", "approval-run", "batch-1", "proposal-1", "approve", "submitted", now),
    )
    storage.execute(
        "INSERT INTO approvals(id,run_id,proposal_id,parsed_action,authorized,status,created_at,consumed_at,proposal_targeting_method) VALUES(?,?,?,?,?,?,?,?,?)",
        ("approval-1", "approval-run", "proposal-1", "approve", 1, "consumed", now, now, "batch"),
    )
    storage.execute(
        "INSERT INTO candidate_risk_budget_decisions(id,run_id,batch_id,candidate_id,proposal_id,symbol,timestamp,passed) VALUES(?,?,?,?,?,?,?,?)",
        ("risk-1", "scan-run", "batch-1", "candidate-1", "proposal-1", "IWM", now, 1),
    )
    storage.execute(
        "INSERT INTO position_sizing_decisions(id,run_id,symbol,timestamp,final_notional,suggested_shares,batch_id,candidate_id,proposal_id) VALUES(?,?,?,?,?,?,?,?,?)",
        ("size-1", "scan-run", "IWM", now, 15, 0.0501, "batch-1", "candidate-1", "proposal-1"),
    )
    storage.execute(
        "INSERT INTO shadow_trades(id,run_id,symbol,side,would_have_entry_price,would_have_entry_time,would_have_notional,reason_not_executed,selected_actual_trade_this_cycle) VALUES(?,?,?,?,?,?,?,?,?)",
        ("shadow-1", "scan-run", "IWM", "buy", 299.08, now, 15, "suppressed", 0),
    )
    storage.execute(
        "INSERT INTO orders(id,run_id,proposal_id,broker_order_id,client_order_id,symbol,side,notional,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        ("order-1", "approval-run", "proposal-1", "broker-1", "client-1", "IWM", "buy", 15, "submitted", now, now),
    )

    broker = ReconcileBroker()
    BrokerReconciler(broker, storage, "reconcile-run").reconcile()

    actual = storage.fetch_all("SELECT * FROM trade_outcomes WHERE actual_or_shadow='actual'")
    assert len(actual) == 1
    row = actual[0]
    assert row["trade_id"] == "order-1"
    assert row["batch_id"] == "batch-1"
    assert row["candidate_id"] == "candidate-1"
    assert row["proposal_id"] == "proposal-1"
    assert row["order_id"] == "order-1"
    assert row["broker_order_id"] == "broker-1"
    assert row["risk_budget_decision_id"] == "risk-1"
    assert row["position_sizing_decision_id"] == "size-1"
    assert row["approval_batch_action_id"] == "batch-action-1"
    assert row["quantity"] == 2
    assert row["entry_price"] == 10.5
    assert row["outcome_status"] == "pending_forward_returns"
    assert row["source"] == "ranked_batch_approval"
    assert storage.fetch_all("SELECT selected_actual_trade_this_cycle, reason_not_executed FROM shadow_trades WHERE id='shadow-1'")[0] == {
        "selected_actual_trade_this_cycle": 1,
        "reason_not_executed": "executed_as_actual",
    }
    assert storage.fetch_all("SELECT order_id, broker_order_id, fill_id FROM candidate_risk_budget_decisions WHERE id='risk-1'")[0]["order_id"] == "order-1"
    assert storage.fetch_all("SELECT candidate_status FROM proposal_batch_candidates WHERE id='candidate-1'")[0]["candidate_status"] == "filled"
    assert broker.submit_calls == 0


def test_batch_candidate_creation_backfills_measurement_linkage(tmp_path):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    now = datetime.now(UTC).isoformat()
    storage.execute(
        "INSERT INTO trade_proposals(id,run_id,symbol,side,notional,status,created_at,expires_at,payload) VALUES(?,?,?,?,?,?,?,?,?)",
        ("proposal-1", "scan-run", "SPY", "buy", 10, "pending", now, (datetime.now(UTC) + timedelta(minutes=5)).isoformat(), "{}"),
    )
    storage.execute(
        "INSERT INTO candidate_risk_budget_decisions(id,run_id,symbol,timestamp,passed) VALUES(?,?,?,?,?)",
        ("risk-1", "scan-run", "SPY", now, 1),
    )
    storage.execute(
        "INSERT INTO position_sizing_decisions(id,run_id,symbol,timestamp,final_notional) VALUES(?,?,?,?,?)",
        ("size-1", "scan-run", "SPY", now, 10),
    )
    storage.execute(
        "INSERT INTO ranked_opportunity_sets(id,run_id,timestamp,symbol,rank,actionable) VALUES(?,?,?,?,?,?)",
        ("rank-1", "scan-run", now, "SPY", 1, 1),
    )

    storage.link_batch_candidate_records("proposal-1", "batch-1", "candidate-1")

    assert storage.fetch_all("SELECT batch_id,candidate_id,proposal_id FROM candidate_risk_budget_decisions WHERE id='risk-1'")[0] == {
        "batch_id": "batch-1",
        "candidate_id": "candidate-1",
        "proposal_id": "proposal-1",
    }
    assert storage.fetch_all("SELECT batch_id,candidate_id,proposal_id FROM position_sizing_decisions WHERE id='size-1'")[0]["candidate_id"] == "candidate-1"
    assert storage.fetch_all("SELECT batch_id,candidate_id,proposal_id FROM ranked_opportunity_sets WHERE id='rank-1'")[0]["proposal_id"] == "proposal-1"


def test_unknown_reconciliation_never_resubmits(tmp_path):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    insert_local_order(storage)
    broker = ReconcileBroker(fail=True)
    result = BrokerReconciler(broker, storage, "run").reconcile()
    assert result.unknown == 1
    assert storage.fetch_all("SELECT status FROM orders WHERE id='local-1'")[0]["status"] == "pending_new"
    assert broker.submit_calls == 0


def test_auto_execution_is_hard_quarantined(tmp_path):
    service, storage = make_service(tmp_path, AuthoritativeBroker())
    service.config["auto_execution_enabled"] = True
    service.config["auto_execution_mode"] = "paper_high_confidence_only"
    assert service._should_auto_execute({"score": 100, "asset_score": 100, "notional": 1}) is False
    assert storage.fetch_all("SELECT count(*) n FROM approvals")[0]["n"] == 0
    assert storage.fetch_all("SELECT count(*) n FROM audit_events WHERE event_type='auto_execution_blocked'")[0]["n"] == 1


def test_live_is_impossible_even_when_yaml_gates_are_true():
    with pytest.raises(PermissionError, match="not supported"):
        AlpacaBroker({"mode": "live", "live_enabled": True, "explicit_live_confirmation": True}, "live-key", "live-secret")


def test_live_block_happens_before_credentials_are_read(monkeypatch):
    monkeypatch.setattr("app.broker_alpaca.get_secret", lambda name: (_ for _ in ()).throw(AssertionError("credential read")))
    with pytest.raises(PermissionError, match="not supported"):
        AlpacaBroker({"mode": "live", "live_enabled": True, "explicit_live_confirmation": True})


def test_lock_inspection_recovers_only_dead_old_owner(tmp_path, monkeypatch):
    lock = tmp_path / "agent.lockdir"
    lock.mkdir()
    (lock / "pid").write_text("123")
    (lock / "started_at_epoch").write_text("1000")
    monkeypatch.setattr(run_lock, "_pid_exists", lambda pid: False)
    assert run_lock.inspect_lock(lock, now=1201).state == "stale"
    assert run_lock.inspect_lock(lock, now=1050).state == "recent_unknown"
    monkeypatch.setattr(run_lock, "_pid_exists", lambda pid: True)
    assert run_lock.inspect_lock(lock, now=99999).state == "active"


def test_excel_redacts_telegram_and_sensitive_payload(tmp_path):
    storage = Storage(tmp_path / "test.db")
    storage.initialize()
    storage.execute(
        "INSERT INTO approvals(id,run_id,proposal_id,sender_id,raw_message,parsed_action,authorized,status,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
        ("a", "run", None, "123456", "my private approval words", "unclear", 1, "rejected", datetime.now(UTC).isoformat()),
    )
    storage.execute(
        "INSERT INTO audit_events(run_id,event_type,actor,detail,created_at) VALUES(?,?,?,?,?)",
        ("run", "test", "system", '{"api_key":"sensitive-value"}', datetime.now(UTC).isoformat()),
    )
    path = export_excel(storage, {"mode": "paper"}, tmp_path / "report.xlsx")
    workbook = load_workbook(path, read_only=True, data_only=True)
    values = "\n".join(str(cell or "") for sheet in workbook for row in sheet.iter_rows(values_only=True) for cell in row)
    assert "my private approval words" not in values
    assert "123456" not in values
    assert "sensitive-value" not in values
    assert "[REDACTED TELEGRAM TEXT]" in values


def test_telegram_update_redacts_by_default():
    update = {"update_id": 1, "message": {"text": "yes buy spy", "from": {"id": 123}, "chat": {"id": 456}}}
    safe = redact_telegram_update(update)
    assert "yes buy spy" not in str(safe)
    assert "123" not in str(safe)
    assert redact_telegram_update(update, include_raw=True) is update


def test_raw_telegram_dump_patterns_are_ignored():
    import subprocess
    root = Path(__file__).resolve().parents[1]
    for name in ("raw_telegram_updates.json", "data/telegram_updates/private.json"):
        result = subprocess.run(["git", "check-ignore", "-q", name], cwd=root)
        assert result.returncode == 0


class ClockBroker:
    def __init__(self, minutes=None, fail=False):
        self.minutes = minutes
        self.fail = fail

    def get_clock(self):
        if self.fail: raise RuntimeError("clock unavailable")
        now = datetime.now(UTC)
        return SimpleNamespace(is_open=True, timestamp=now, next_close=now + timedelta(minutes=self.minutes))


def expiry_service(broker):
    service = TradingService.__new__(TradingService)
    service.config = load_config()
    service.broker = broker
    return service


def test_near_close_expiry_truncates_and_respects_bounds():
    signal = Signal("ENTRY", "buy", "SPY", "reason", 0.8, {"volatility_20": 0.15})
    now = datetime.now(UTC)
    assert expiry_service(ClockBroker(8))._calculate_expiry_minutes("SPY", signal, 0.15, 70, now, now) == 8
    assert expiry_service(ClockBroker(2))._calculate_expiry_minutes("SPY", signal, 0.15, 70, now, now) == 5
    assert expiry_service(ClockBroker(fail=True))._calculate_expiry_minutes("SPY", signal, 0.15, 70, now, now) == 5
