import os
import json
from datetime import datetime, UTC, timedelta
import pytest
import pandas as pd
import numpy as np

from app.utils import load_config, PROJECT_ROOT
from app.storage import Storage
from app.service import TradingService
from app.strategy_rule_based import Signal, evaluate_symbol
from test_scoring_and_throttling import MockBroker, MockTelegramBot, temp_storage
from app.reports import export_excel
from openpyxl import load_workbook


def test_volatility_score_grading(temp_storage):
    # Verify the Trade Decision Score's volatility component is graded correctly:
    # 8%–25%: 15/15
    # 25%–35%: 10/15
    # 35%–45%: 5/15
    # 0%–8%: 8/15
    # >45%: 0/15
    # missing/invalid: 0/15
    config = load_config()
    broker = MockBroker()
    service = TradingService(config, temp_storage, broker, "test_run_id")
    service.telegram = MockTelegramBot()

    # We mock evaluate_symbol to return signals with different volatility levels
    from app.strategy_rule_based import Signal

    def get_score_for_vol(vol_val):
        import app.service
        original_evaluate = app.service.evaluate_symbol
        def mock_evaluate(*args, **kwargs):
            # To get an ENTRY signal, trend conditions must pass
            return Signal("ENTRY", "buy", "SPY", "Reason", 0.7, {"volatility_20": vol_val})

        app.service.evaluate_symbol = mock_evaluate
        try:
            temp_storage.execute("DELETE FROM orders")
            temp_storage.execute("DELETE FROM fills")
            temp_storage.execute("DELETE FROM approvals")
            temp_storage.execute("DELETE FROM proposal_batches")
            temp_storage.execute("DELETE FROM proposal_batch_candidates")
            temp_storage.execute("DELETE FROM trade_proposals")
            temp_storage.execute("DELETE FROM market_memory")
            service.scan()
        finally:
            app.service.evaluate_symbol = original_evaluate

        rows = temp_storage.fetch_all("SELECT * FROM market_memory ORDER BY created_at DESC LIMIT 1")
        if not rows:
            return None, None
        return rows[0]["volatility_score_contribution"], rows[0]["volatility_regime"]

    # 1. 0%-8% (e.g. 0.05) -> 8/15, "too quiet"
    score, regime = get_score_for_vol(0.05)
    assert score == 8.0
    assert regime == "too quiet"

    # 2. 8%-25% (e.g. 0.15) -> 15/15, "normal"
    score, regime = get_score_for_vol(0.15)
    assert score == 15.0
    assert regime == "normal"

    # 3. 25%-35% (e.g. 0.30) -> 10/15, "elevated"
    score, regime = get_score_for_vol(0.30)
    assert score == 10.0
    assert regime == "elevated"

    # 4. 35%-45% (e.g. 0.40) -> 5/15, "high"
    score, regime = get_score_for_vol(0.40)
    assert score == 5.0
    assert regime == "high"

    # 5. >45% (e.g. 0.50) -> 0/15, "extreme"
    score, regime = get_score_for_vol(0.50)
    assert score == 0.0
    assert regime == "extreme"

    # 6. missing/None -> 0/15, "missing"
    score, regime = get_score_for_vol(None)
    assert score == 0.0
    assert regime == "missing"


def test_volatility_regime_eligibility_and_size_adjustments(temp_storage):
    # Verify new entries and notional adjustments:
    # 8%–25%: eligible, normal paper size
    # 25%–35%: eligible, 50% paper size
    # 35%–45%: watch-only, no proposal generated
    # >45%: extreme blocked, no proposal generated
    config = load_config()
    config["phase3"]["enabled"] = False
    config["phase3"]["active"] = False
    config["phase4"]["enabled"] = False
    config["phase4"]["active"] = False
    config.setdefault("risk", {})["require_gpt_review_for_buy_proposals"] = False
    broker = MockBroker()
    service = TradingService(config, temp_storage, broker, "test_run_id")
    service.telegram = MockTelegramBot()

    def run_scan_with_vol(vol_val):
        from app.strategy_rule_based import Signal
        import app.service
        original_evaluate = app.service.evaluate_symbol
        def mock_evaluate(symbol, *args, **kwargs):
            if symbol != "SPY":
                return Signal("HOLD", None, symbol, "not SPY", 0.0, {})
            if vol_val is None:
                return Signal("HOLD", None, "SPY", "missing volatility data; fail-safe HOLD", 0.0, {})
            elif vol_val > 0.45:
                return Signal("HOLD", None, "SPY", "extreme volatility; blocked", 0.0, {"volatility_20": vol_val})
            elif vol_val > 0.35:
                return Signal("HOLD", None, "SPY", "high volatility; watch only", 0.0, {"volatility_20": vol_val})
            elif vol_val >= 0.25:
                return Signal("ENTRY", "buy", "SPY", "trend filters passed and volatility elevated; reduced confidence", 0.7, {"volatility_20": vol_val})
            else:
                return Signal("ENTRY", "buy", "SPY", "trend filters passed and volatility normal", 0.7, {"volatility_20": vol_val})

        app.service.evaluate_symbol = mock_evaluate
        try:
            temp_storage.execute("DELETE FROM orders")
            temp_storage.execute("DELETE FROM fills")
            temp_storage.execute("DELETE FROM approvals")
            temp_storage.execute("DELETE FROM proposal_batches")
            temp_storage.execute("DELETE FROM proposal_batch_candidates")
            temp_storage.execute("DELETE FROM trade_proposals")
            temp_storage.execute("DELETE FROM market_memory")
            service.scan()
        finally:
            app.service.evaluate_symbol = original_evaluate

    # Case A: 15% (normal volatility) -> eligible, normal notional
    run_scan_with_vol(0.15)
    props = temp_storage.fetch_all("SELECT * FROM trade_proposals WHERE symbol='SPY'")
    assert len(props) == 1
    p_payload = json.loads(props[0]["payload"])
    assert p_payload["notional"] > 250.0
    normal_notional = p_payload["notional"]
    assert "stage" not in p_payload["sizing_caps"]
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY'")[0]
    assert memory["volatility_regime"] == "normal"
    assert memory["paper_size_adjustment"] == 1.0

    # Case B: 30% (elevated volatility) -> eligible, 50% notional
    run_scan_with_vol(0.30)
    props = temp_storage.fetch_all("SELECT * FROM trade_proposals WHERE symbol='SPY'")
    assert len(props) == 1
    p_payload = json.loads(props[0]["payload"])
    assert 0 < p_payload["notional"] <= normal_notional
    assert "stage" not in p_payload["sizing_caps"]
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY'")[0]
    assert memory["volatility_regime"] == "elevated"
    assert memory["paper_size_adjustment"] == 0.5

    # Case C: 40% (high volatility) -> watch-only, no proposal
    run_scan_with_vol(0.40)
    props = temp_storage.fetch_all("SELECT * FROM trade_proposals WHERE symbol='SPY'")
    assert len(props) == 0
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY'")[0]
    assert memory["volatility_regime"] == "high"
    assert memory["volatility_gate_result"] == "watch only"
    assert memory["proposal_generated"] == 0

    # Case D: 50% (extreme volatility) -> blocked, no proposal
    run_scan_with_vol(0.50)
    props = temp_storage.fetch_all("SELECT * FROM trade_proposals WHERE symbol='SPY'")
    assert len(props) == 0
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY'")[0]
    assert memory["volatility_regime"] == "extreme"
    assert memory["volatility_gate_result"] == "blocked"
    assert memory["proposal_generated"] == 0


def test_state_based_proposal_deduplication(temp_storage):
    config = load_config()
    config["phase3"]["enabled"] = False
    config["phase3"]["active"] = False
    config["phase4"]["enabled"] = False
    config["phase4"]["active"] = False
    config.setdefault("risk", {})["require_gpt_review_for_buy_proposals"] = False
    broker = MockBroker()
    service = TradingService(config, temp_storage, broker, "test_run_id")
    service.telegram = MockTelegramBot()

    def run_scan_with_custom_signal(action, side, score, symbol="SPY", vol=0.15):
        from app.strategy_rule_based import Signal
        import app.service
        original_evaluate = app.service.evaluate_symbol
        def mock_evaluate(sym, *args, **kwargs):
            if sym != symbol:
                return Signal("HOLD", None, sym, "not target symbol", 0.0, {})
            return Signal(action, side, symbol, "Reason", 0.7, {"volatility_20": vol})

        original_score = service._calculate_asset_selection_score
        service._calculate_asset_selection_score = lambda *a, **k: 80.0

        app.service.evaluate_symbol = mock_evaluate
        try:
            service.scan()
        finally:
            app.service.evaluate_symbol = original_evaluate
            service._calculate_asset_selection_score = original_score

    now = datetime.now(UTC)

    # 1. Pending proposal blocks duplicate
    temp_storage.execute("DELETE FROM trade_proposals")
    temp_storage.execute("DELETE FROM market_memory")
    temp_storage.execute(
        "INSERT INTO trade_proposals(id,run_id,signal_id,symbol,side,notional,status,created_at,expires_at,strategy_version,payload) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        ("p1", "r1", "s1", "SPY", "buy", 5, "pending", now.isoformat(), (now + timedelta(minutes=10)).isoformat(), "rule_based_v1", json.dumps({"score": 70}))
    )

    run_scan_with_custom_signal("ENTRY", "buy", 70)
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY' ORDER BY created_at DESC LIMIT 1")[0]
    assert memory["dedupe_status"] == "suppressed"
    assert "active/pending similar proposal exists" in memory["dedupe_reason"]
    assert memory["proposal_generated"] == 0

    # 2. Cooldown blocks duplicate within 60 minutes
    temp_storage.execute("DELETE FROM trade_proposals")
    temp_storage.execute("DELETE FROM market_memory")
    created_at = now - timedelta(minutes=10)
    temp_storage.execute(
        "INSERT INTO trade_proposals(id,run_id,signal_id,symbol,side,notional,status,created_at,expires_at,strategy_version,payload,setup_key) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
        ("p2", "r1", "s2", "SPY", "buy", 5, "approved", created_at.isoformat(), (created_at + timedelta(minutes=10)).isoformat(), "rule_based_v1", json.dumps({"symbol": "SPY", "score": 95.0, "notional": 5.0, "latest_price": 100.0, "stop_distance_dollars": 10.0, "stop_risk_dollars": 0.5, "cluster_name": "us_broad_market"}), "SPY:buy:ENTRY:below_50:above_200:normal:score_90")
    )

    run_scan_with_custom_signal("ENTRY", "buy", 75)
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY' ORDER BY created_at DESC LIMIT 1")[0]
    assert memory["dedupe_status"] == "suppressed"
    assert "duplicate proposal cooldown" in memory["dedupe_reason"]
    assert memory["proposal_generated"] == 0

    # 3. Meaningful score improvement (delta >= 10) allows new proposal
    temp_storage.execute("DELETE FROM trade_proposals")
    temp_storage.execute("DELETE FROM market_memory")
    temp_storage.execute(
        "INSERT INTO trade_proposals(id,run_id,signal_id,symbol,side,notional,status,created_at,expires_at,strategy_version,payload,setup_key) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
        ("p3", "r1", "s3", "SPY", "buy", 5, "approved", created_at.isoformat(), (created_at + timedelta(minutes=10)).isoformat(), "rule_based_v1", json.dumps({"symbol": "SPY", "score": 70.0, "notional": 5.0, "latest_price": 100.0, "stop_distance_dollars": 10.0, "stop_risk_dollars": 0.5, "cluster_name": "us_broad_market"}), "SPY:buy:ENTRY:below_50:above_200:normal:score_90")
    )

    # Run scan which will result in score 80 (delta = 10 from 70)
    run_scan_with_custom_signal("ENTRY", "buy", 80)
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY' ORDER BY created_at DESC LIMIT 1")[0]
    assert memory["dedupe_status"] == "allowed"
    assert "meaningful score improvement" in memory["dedupe_reason"]
    assert memory["proposal_generated"] == 1

    # 4. Exit/reduce-risk action is not blocked if there is an active position
    temp_storage.execute("DELETE FROM trade_proposals")
    temp_storage.execute("DELETE FROM market_memory")
    temp_storage.execute(
        "INSERT INTO trade_proposals(id,run_id,signal_id,symbol,side,notional,status,created_at,expires_at,strategy_version,payload,setup_key) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
        ("p4", "r1", "s4", "SPY", "sell", 5, "approved", (now - timedelta(minutes=5)).isoformat(), now.isoformat(), "rule_based_v1", json.dumps({"score": 95.0}), "SPY:sell:EXIT:below_50:above_200:normal:score_90")
    )
    broker.positions = [type("Pos", (), {"symbol": "SPY", "qty": 10, "market_value": 5000})()]

    run_scan_with_custom_signal("EXIT", "sell", 70)
    memory = temp_storage.fetch_all("SELECT * FROM market_memory WHERE symbol='SPY' ORDER BY created_at DESC LIMIT 1")[0]
    assert memory["dedupe_status"] == "allowed"
    assert "exit/reduce-risk action" in memory["dedupe_reason"]
    assert memory["proposal_generated"] == 1


def test_excel_export_includes_volatility_and_dedupe_fields(temp_storage, tmp_path):
    config = load_config()
    broker = MockBroker()
    service = TradingService(config, temp_storage, broker, "test_run_id")
    service.telegram = MockTelegramBot()

    temp_storage.execute("DELETE FROM market_memory")
    temp_storage.execute(
        "INSERT INTO market_memory(run_id,market_profile,symbol,price,prev_price,price_change,price_change_pct,session_start_price,session_change,volatility,signal,score,classification,reason,proposal_allowed,gpt_called,created_at,asset_score,asset_classification,symbol_rank,proposal_generated,no_action_reason,asset_selection_score,trade_decision_score,system_confidence,gpt_confidence,gpt_caution,expiry_minutes,expires_at_sgt,main_risk,volatility_regime,volatility_score_contribution,volatility_gate_result,dedupe_status,dedupe_reason,paper_size_adjustment) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        ("run1", "default", "SPY", 500.0, 499.0, 1.0, 0.2, 500.0, 0.0, 0.15, "ENTRY", 85.0, "Strong", "trend", 1, 0, "2026-06-22T17:00:00", 80.0, "Strong", 1, 1, "generated", 80.0, 85.0, "Strong", "N/A", "N/A", 15, "SGT", "N/A", "normal", 15.0, "eligible", "allowed", "passed check", 1.0)
    )

    excel_path = tmp_path / "test_report_vol.xlsx"
    export_excel(temp_storage, config, excel_path)

    workbook = load_workbook(excel_path)
    sheet = workbook["Market Memory"]
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) > 1
    headers = rows[0]

    assert "volatility_regime" in headers
    assert "volatility_score_contribution" in headers
    assert "volatility_gate_result" in headers
    assert "dedupe_status" in headers
    assert "dedupe_reason" in headers
    assert "paper_size_adjustment" in headers

    data_row = rows[1]
    headers_indices = {h: idx for idx, h in enumerate(headers)}
    assert data_row[headers_indices["volatility_regime"]] == "normal"
    assert data_row[headers_indices["volatility_score_contribution"]] == 15.0
    assert data_row[headers_indices["volatility_gate_result"]] == "eligible"
    assert data_row[headers_indices["dedupe_status"]] == "allowed"
    assert data_row[headers_indices["dedupe_reason"]] == "passed check"
    assert data_row[headers_indices["paper_size_adjustment"]] == 1.0
